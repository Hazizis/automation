from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import time

wb = load_workbook(filename="D:\Autmation\data.xlsx")

sheetRange = wb.active

total_rows = sheetRange.max_row
total_col = sheetRange.max_column


#Nik = (sheetRange.cell(row=3, column=3))
#print (Nik.value)

i=2

while i <= len(sheetRange['A']):
    Nik = sheetRange['C'+str(i)].value

    try:
        print (Nik)
    except TimeoutException:
        print ("data tidak masuk")
        pass
    i = i+1
