from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
import time

options = Options()
options.add_experimental_option("detach", True)

wb = load_workbook(filename="D:\Autmation\data.xlsx")

sheetRange = wb.active


driver = webdriver.Chrome(options=options)
url1='https://subsiditepatlpg.mypertamina.id/merchant/auth/login'
url2='https://subsiditepatlpg.mypertamina.id/merchant/app/verification-nik'
driver.maximize_window()
driver.implicitly_wait(10)
driver.get(url2)
driver.find_element(By.ID, "mantine-r0").send_keys("085337580232")
time.sleep(2)
driver.find_element(By.ID, "mantine-r1").send_keys("123456")
time.sleep(1)
driver.find_element(By.CLASS_NAME, "styles_btnLogin__wsKTT").click()
time.sleep(2)
driver.find_element(By.CLASS_NAME, "styles_iconClose__ZjGFM").click()
#driver.get(url2)
Nik = sheetRange.cell(row=3, column=3)

wait = WebDriverWait(driver, 20)
typetextfirst = wait.until(EC.element_to_be_clickable((By.ID, "mantine-rd")))
typetextfirst.click()
typetextfirst.send_keys(Nik.value)
driver.find_element(By.CLASS_NAME, "styles_headerForm__t7P4g").click()
#driver.find_element(By.CLASS_NAME, "mantine-Input-input mantine-Autocomplete-input mantine-1w8ojc8").send_keys("5203024804700001")
time.sleep(2)
driver.find_element(By.CLASS_NAME, "styles_btnBayar__o4O4A").click()




'''
i=2

while i <= len(sheetRange['C']):
    Nik = sheetRange['C'+str(i)].value

    try:
        driver.find_element(By.ID, "mantine-ra").send_keys(Nik)
        time.sleep(5)
        driver.find_element(By.ID, "btnCheckNik").click()

    except TimeoutException:
        print ("data tidak masuk")
        pass

    time.sleep(5)
'''
